"""
Excel integration module for handling Excel file operations.
"""
from typing import List, Dict, Any, Optional
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from fastapi import HTTPException, UploadFile

from app.models.order import Order
from app.models.customer import Customer
from app.models.item import Item
from app.schemas.order import OrderCreate

def process_order_excel(file: UploadFile, db: Any) -> List[OrderCreate]:
    """
    Excelファイルから注文データを処理する関数
    
    Args:
        file (UploadFile): アップロードされたExcelファイル
        db (Session): データベースセッション
    
    Returns:
        List[OrderCreate]: 処理された注文データのリスト
    
    Raises:
        HTTPException: Excelファイルの処理中にエラーが発生した場合
    """
    try:
        # Excelファイルを読み込む
        workbook = openpyxl.load_workbook(file.file)
        sheet = workbook.active
        
        orders: List[OrderCreate] = []
        # ヘッダー行をスキップ
        for row in sheet.iter_rows(min_row=2):
            # 必要なデータが存在する場合のみ処理
            if row[0].value and row[1].value and row[2].value:
                customer_code = str(row[0].value)
                item_code = str(row[1].value)
                quantity = int(row[2].value)
                
                # 顧客とアイテムの存在確認
                customer = db.query(Customer).filter(Customer.code == customer_code).first()
                item = db.query(Item).filter(Item.code == item_code).first()
                
                if not customer:
                    raise HTTPException(
                        status_code=400,
                        detail=f"顧客コード {customer_code} が見つかりません"
                    )
                if not item:
                    raise HTTPException(
                        status_code=400,
                        detail=f"商品コード {item_code} が見つかりません"
                    )
                
                # 注文データの作成
                order = OrderCreate(
                    customer_id=customer.id,
                    item_id=item.id,
                    quantity=quantity,
                    order_date=datetime.now()
                )
                orders.append(order)
        
        return orders
    
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"Excelファイルの処理中にエラーが発生しました: {str(e)}"
        )

def generate_order_template() -> bytes:
    """
    注文用のExcelテンプレートを生成する関数
    
    Returns:
        bytes: 生成されたExcelファイルのバイトデータ
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "注文データ"
    
    # ヘッダーの設定
    headers = ["顧客コード", "商品コード", "数量", "備考"]
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # 列幅の設定
    for col in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 15
    
    # ファイルをバイトストリームとして保存
    from io import BytesIO
    excel_file = BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)
    
    return excel_file.read() 